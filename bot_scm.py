import time
import requests
import json
import pandas as pd
import io
import os
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def jalankan_bot():
    print("=== [LOG START] Memulai Operasi Bot SCM ===")
    
    # 1. Konfigurasi Browser Siluman (OTOMATIS)
    options = uc.ChromeOptions()
    options.add_argument('--headless') # Wajib di GitHub Actions
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    
    # User agent agar terlihat seperti browser asli manusia
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

    print("[1/6] Menyiapkan Browser Siluman (Auto-Detecting Chrome)...")
    
    try:
        # Menghapus 'version_main' agar script mencari versi Chrome terbaru sendiri
        driver = uc.Chrome(options=options)
        
        print("[2/6] Membuka Halaman Login: https://scm.nusadaya.net/login")
        driver.get("https://scm.nusadaya.net/login")
        
        wait = WebDriverWait(driver, 25)
        
        print("[3/6] Mengisi Form Login...")
        email_rahasia = os.environ.get('EMAIL_SCM')
        pass_rahasia = os.environ.get('PASS_SCM')

        if not email_rahasia or not pass_rahasia:
            print("(!) Error: Secrets EMAIL_SCM atau PASS_SCM belum diatur di GitHub!")
            return

        email_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @placeholder='Email atau NIP']")))
        email_input.send_keys(email_rahasia)
        
        pass_input = driver.find_element(By.XPATH, "//input[@type='password']")
        pass_input.send_keys(pass_rahasia)
        
        print("-> Mengetuk tombol Log in...")
        login_btn = driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]")
        login_btn.click()
        
        print("-> Login dikirim, menunggu dashboard (15 detik)...")
        time.sleep(15)
        
        cookies = driver.get_cookies()
        if not cookies:
            print("(!) Peringatan: Tidak ada cookie terdeteksi. Login mungkin gagal.")
        else:
            print(f"-> Berhasil mendapatkan {len(cookies)} cookies.")

        print("[4/6] Menembak URL Export Langsung...")
        session_cookies = {c['name']: c['value'] for c in cookies}
        export_url = "https://scm.nusadaya.net/purchase-order/export?tahun=2026&up=all&bidang=all&khs=all"
        
        response_dl = requests.get(export_url, cookies=session_cookies)
        
        if response_dl.status_code == 200:
            print(f"-> Download Sukses! Ukuran file: {len(response_dl.content)} bytes")
        else:
            print(f"(!) Gagal download. Status Code: {response_dl.status_code}")
            return

        print("[5/6] Membongkar isi Excel & Mengekstrak Hyperlink...")
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=False)
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(min_row=2):
            current_row = []
            for cell in row:
                if cell.hyperlink:
                    label = str(cell.value) if cell.value is not None else "Lihat File"
                    url = cell.hyperlink.target
                    current_row.append(f"{label} {url}")
                else:
                    current_row.append(cell.value if cell.value is not None else "")
            data_rows.append(current_row)

        print(f"-> Total data ditemukan: {len(data_rows)} baris.")

        print("[6/6] Mengirim data ke Google Sheets...")
        gas_url = "https://script.google.com/macros/s/AKfycbzrpSwYWqzvtvqugpWI2UTr6Ivg2C87qFkIO4UrYwyaBIglD8zX7jDV_aPNzBImwSI/exec"
        
        payload = json.dumps({"rows": data_rows})
        res_gas = requests.post(gas_url, data=payload, headers={'Content-Type': 'application/json'})
        
        print(f"=== [HASIL AKHIR] Respon dari Google Sheets: {res_gas.text} ===")

    except Exception as e:
        print(f"!!! [ERROR] Terjadi kendala teknis: {str(e)}")
    finally:
        print("=== [LOG END] Menutup Browser & Selesai ===")
        try:
            driver.quit()
        except:
            pass

if __name__ == "__main__":
    jalankan_bot()
